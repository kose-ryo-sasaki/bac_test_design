import streamlit as st
import pandas as pd
import matplotlib.colors as mcolors
import seaborn as sns
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from datetime import datetime
import os
from dotenv import load_dotenv


# **ローカルでは `.env` を読み込む**
if os.path.exists(".env"):
    load_dotenv()
    PASSWORD = os.getenv("PASSWORD")
else:
    # **Cloud では Streamlit Secrets から取得**
    PASSWORD = st.secrets.get("PASSWORD")

# **Secrets にパスワードが設定されていない場合の処理**
if PASSWORD is None:
    st.error("🔐 パスワードが設定されていません！")
    st.stop()

# **セッションにログイン情報がない場合は初期化**
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

# **ログイン画面**
if not st.session_state["authenticated"]:
    st.title("ログイン画面")
    password_input = st.text_input("パスワードを入力してください:", type="password")

    if st.button("ログイン"):
        if password_input == PASSWORD:  # **一致すればログイン成功**
            st.session_state["authenticated"] = True
            st.rerun()  # **ログイン後にリロード**
        else:
            st.error("パスワードが間違っています。")

    st.stop()  # **ログイン成功しない限りアプリを進めない**


# **セッションにログイン情報がない場合は初期化**
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

# **ログイン画面**
if not st.session_state["authenticated"]:
    st.title("🔒 ログインが必要です")
    password_input = st.text_input("パスワードを入力してください:", type="password")

    if st.button("ログイン"):
        if password_input == PASSWORD:  # **一致すればログイン成功**
            st.session_state["authenticated"] = True
            st.rerun()  # **ログイン後にリロード**
        else:
            st.error("パスワードが間違っています。")

    st.stop()  # **ログイン成功しない限りアプリを進めない**

# **ページの設定**
st.set_page_config(layout="wide")
st.title("9×12 配置エディター")

# **CSVファイルのアップロード**
uploaded_file = st.file_uploader("CSVファイルをアップロードしてください", type=["csv"])

if uploaded_file is not None:
    # **CSVを読み込む**
    df = pd.read_csv(uploaded_file)

    # **iter の数だけデータを展開**
    df_expanded = df.loc[df.index.repeat(df['iter'])].copy()
    df_expanded['iter_count'] = df_expanded.groupby(['sample_id', 'sample_name', 'bac', 'conc']).cumcount() + 1
    df_expanded = df_expanded.drop(columns=['iter'])

    # **行ラベル（A～J）と列ラベル（1～12）を定義**
    row_labels = list("ABCDEFGHIJ")
    col_labels = [str(i) for i in range(1, 13)]

    # **A1, A2, ... のように順番に position を割り当てる**
    positions = [f"{row}{col}" for row in row_labels for col in col_labels]
    df_expanded["position"] = positions[:len(df_expanded)]

    # **元データ（編集不可）を表示**
    st.write("### 元データ（編集不可）")
    st.dataframe(df_expanded)

    # **9×12の表を作成**
    reshaped_df = pd.DataFrame("", index=row_labels, columns=col_labels)

    for _, row in df_expanded.iterrows():
        reshaped_df.loc[row["position"][0], row["position"][1:]] = f"{row['sample_name']}_{row['bac']}_{row['conc']}_{row['iter_count']}"

    # **9×12の表を編集可能にして表示**
    st.write("### 9×12 試験データ表（編集可能）")
    edited_reshaped_df = st.data_editor(reshaped_df, num_rows="dynamic")

    # **位置情報の更新辞書（bac を含めたキーで管理）**
    new_position_mapping = {}

    for row_label in edited_reshaped_df.index:
        for col_label in edited_reshaped_df.columns:
            cell_value = edited_reshaped_df.loc[row_label, col_label]
            if cell_value:
                sample_name, bac, conc, iter_count = cell_value.split("_")
                key = f"{sample_name}_{bac}_{conc}_{iter_count}"

                if key not in new_position_mapping:
                    new_position_mapping[key] = []
                new_position_mapping[key].append(f"{row_label}{col_label}")

    df_expanded["unique_key"] = df_expanded.apply(lambda row: f"{row['sample_name']}_{row['bac']}_{row['conc']}_{row['iter_count']}", axis=1)

    df_updated = df_expanded.copy()

    used_positions = set()
    def assign_position(row):
        key = f"{row['sample_name']}_{row['bac']}_{row['conc']}_{row['iter_count']}"
        possible_positions = new_position_mapping.get(key, [])

        for pos in possible_positions:
            if pos not in used_positions:
                used_positions.add(pos)
                return pos

        return row["position"]

    df_updated["position"] = df_updated.apply(assign_position, axis=1)

    df_updated = df_updated.drop(columns=["unique_key"])

    # **表示用のデータ（bac を除外）**
    df_updated["display_value"] = df_updated.apply(
        lambda row: f"{row['sample_name']}_{row['conc']}_{row['iter_count']}",
        axis=1
    )

    # **色付きの9×12表を作成**
    color_reshaped_df = pd.DataFrame("", index=row_labels, columns=col_labels)
    color_mapping = {}

    # `bac` ごとに異なる色を割り当てる
    unique_bacs = df_updated["bac"].unique()
    color_palette = sns.color_palette("pastel", len(unique_bacs))
    color_map = {bac: mcolors.to_hex(color_palette[i]) for i, bac in enumerate(unique_bacs)}

    # **色付き表のデータを作成**
    for _, row in df_updated.iterrows():
        pos = row["position"]
        row_label, col_label = pos[0], pos[1:]
        color_reshaped_df.loc[row_label, col_label] = row["display_value"]
        color_mapping[(row_label, col_label)] = color_map[row["bac"]]  # `position` をキーに `bac` の色を適用

    # **色を適用する関数**
    def apply_color(data):
        """各セルの position に対応する色を適用"""
        styles = pd.DataFrame("", index=data.index, columns=data.columns)
        for row_label in data.index:
            for col_label in data.columns:
                if (row_label, col_label) in color_mapping:
                    styles.loc[row_label, col_label] = f'background-color: {color_mapping[(row_label, col_label)]}; color: black;'
        return styles

    # **凡例を先に表示**
    st.write("### 凡例（bac ごとの色）")
    legend_html = "".join(
        [f'<div style="display: inline-block; width: 20px; height: 20px; background-color: {color}; margin-right: 5px;"></div> {bac}'
        for bac, color in color_map.items()]
    )
    st.markdown(legend_html, unsafe_allow_html=True)

    # **色付きの表を表示**
    st.write("### 色付き9×12試験データ表")

    styled_df = color_reshaped_df.style.apply(apply_color, axis=None)  # axis=Noneで全体に適用
    st.table(styled_df)

    st.write("### 編集後元データ")
    st.dataframe(df_updated)
    
    # **Excelファイルを作成**
    def create_excel_file(df, color_mapping, file_name):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "9x12_Table"

        # **1行目にファイル名と日付**
        ws["A1"] = f"File: {file_name}"
        ws["B1"] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # **B2セルから列名 (1,2,3,...) を追加**
        for j, col_name in enumerate(df.columns, start=2):  # B2, C2, D2...
            ws.cell(row=2, column=j, value=col_name)

        # **A3セルから行名 (A,B,C,...) を追加**
        for i, row_name in enumerate(df.index, start=3):  # A3, A4, A5...
            ws.cell(row=i, column=1, value=row_name)

        # **B3セルからデータを入力**
        for i, row in enumerate(df.index):
            for j, col in enumerate(df.columns):
                cell = ws.cell(row=i + 3, column=j + 2, value=df.loc[row, col])

                # **色を適用**
                position = (row, col)  # (行ラベル, 列ラベル)
                if position in color_mapping:
                    fill_color = color_mapping[position][1:]  # `#RRGGBB` から `RRGGBB` に変換
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                # **罫線を適用**
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
                cell.border = thin_border

        # **右側に凡例を表示**
        ws["O2"] = "Legend"
        for i, (bac, color) in enumerate(color_map.items()):
            ws.cell(row=i + 3, column=15, value=bac)
            ws.cell(row=i + 3, column=16).fill = PatternFill(
                start_color=color[1:], end_color=color[1:], fill_type="solid"
            )

        # **バイナリデータとして保存**
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # **df_updated を CSV でダウンロード**
    def create_csv_file(df):
        output = BytesIO()
        df.to_csv(output, index=False, encoding="utf-8-sig")
        output.seek(0)  # **バッファの先頭に戻る**
        return output

    st.download_button(
        label="更新後のデータ (CSV) をダウンロード",
        data=create_csv_file(df_updated),
        file_name="updated_data.csv",
        mime="text/csv"
    )

    # **9×12の色付きデータを Excel でダウンロード**
    st.download_button(
        label="色付きExcelをダウンロード",
        data=create_excel_file(color_reshaped_df, color_mapping, "9x12_table.xlsx"),
        file_name="9x12_table.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )